#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import json
import os
import re
import zipfile
import xml.etree.ElementTree as ET
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


def slugify(name: str) -> str:
    s = re.sub(r"\s+", "_", name.strip())
    s = re.sub(r"[^0-9A-Za-z_\u4e00-\u9fff-]+", "", s)
    return s[:80] or "item"


def parse_min_order(min_order: Optional[str]) -> Optional[float]:
    if not min_order:
        return None
    m = re.search(r"(\d+(?:\.\d+)?)", str(min_order))
    return float(m.group(1)) if m else None


def parse_dispimg_id(cell_value: Any) -> Optional[str]:
    if not isinstance(cell_value, str):
        return None
    m = re.search(r'DISPIMG\("(?P<id>ID_[0-9A-Fa-f]+)"', cell_value)
    return m.group("id") if m else None


def build_id_to_media_map(xlsx_path: str) -> Dict[str, str]:
    """
    Map DISPIMG ID_* to xlsx internal media path (like xl/media/image81.jpeg).
    This file uses WPS cellImages extension: xl/cellimages.xml + xl/_rels/cellimages.xml.rels
    """
    ns = {
        "wps": "http://www.wps.cn/officeDocument/2017/etCustomData",
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    with zipfile.ZipFile(xlsx_path) as z:
        cell_xml = z.read("xl/cellimages.xml")
        rels_xml = z.read("xl/_rels/cellimages.xml.rels")

    rels_root = ET.fromstring(rels_xml)
    relmap = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels_root.findall(".//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship")
    }

    root = ET.fromstring(cell_xml)
    out: Dict[str, str] = {}
    for cell_image in root.findall("wps:cellImage", ns):
        c_nv_pr = cell_image.find(".//xdr:cNvPr", ns)
        if c_nv_pr is None:
            continue
        img_id = c_nv_pr.attrib.get("name")
        if not img_id or not img_id.startswith("ID_"):
            continue
        blip = cell_image.find(".//a:blip", ns)
        if blip is None:
            continue
        rid = blip.attrib.get("{%s}embed" % ns["r"])
        if not rid:
            continue
        target = relmap.get(rid)
        if not target:
            continue
        # target like "media/image81.jpeg" (relative to xl/)
        out[img_id] = f"xl/{target}"
    return out


def build_cell_to_media_map_from_drawing(xlsx_path: str) -> Dict[Tuple[int, int], str]:
    """
    Map (row_1based, col_1based) to xlsx internal media path (like xl/media/image2.jpeg)
    using standard drawing anchors (xl/drawings/drawing*.xml).
    """
    ns = {
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    with zipfile.ZipFile(xlsx_path) as z:
        # Most files use drawing1.xml for the first sheet. If missing, skip.
        drawing_xml_name = "xl/drawings/drawing1.xml"
        rels_xml_name = "xl/drawings/_rels/drawing1.xml.rels"
        if drawing_xml_name not in z.namelist() or rels_xml_name not in z.namelist():
            return {}
        drawing_xml = z.read(drawing_xml_name)
        rels_xml = z.read(rels_xml_name)

    rels_root = ET.fromstring(rels_xml)
    relmap = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels_root.findall(".//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship")
    }
    root = ET.fromstring(drawing_xml)

    anchors = root.findall("xdr:twoCellAnchor", ns) + root.findall("xdr:oneCellAnchor", ns)
    out: Dict[Tuple[int, int], str] = {}
    for a in anchors:
        fr = a.find("xdr:from", ns)
        if fr is None:
            continue
        r0 = fr.find("xdr:row", ns)
        c0 = fr.find("xdr:col", ns)
        if r0 is None or c0 is None:
            continue
        # Drawing anchor row/col are 0-based indexes.
        row_1 = int(r0.text) + 1
        col_1 = int(c0.text) + 1

        blip = a.find(".//a:blip", ns)
        if blip is None:
            continue
        rid = blip.attrib.get("{%s}embed" % ns["r"])
        if not rid:
            continue
        target = relmap.get(rid)
        if not target:
            continue
        # target like "../media/image2.jpeg" (relative to xl/drawings/)
        # normalize to "xl/media/..."
        target = target.replace("\\", "/")
        if target.startswith("../"):
            target = target[3:]
        out[(row_1, col_1)] = f"xl/{target}"
    return out


def is_category_row(row_values: List[Any]) -> Optional[str]:
    a = row_values[0]
    if not isinstance(a, str):
        return None
    title = a.strip()
    if not title:
        return None
    # Heuristic: category row has only A filled (or mostly empty)
    non_empty = sum(1 for v in row_values if v is not None and str(v).strip())
    if non_empty <= 2 and title.startswith("当夏"):
        return title
    return None


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--out-json", required=True)
    ap.add_argument("--assets-dir", required=True)
    args = ap.parse_args()

    os.makedirs(args.assets_dir, exist_ok=True)

    # Prefer standard drawing anchors (more complete).
    cell_to_media = build_cell_to_media_map_from_drawing(args.xlsx)
    id_to_media = build_id_to_media_map(args.xlsx)

    wb = load_workbook(args.xlsx, data_only=False)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]

    # Store image paths relative to the web root (dessert-quote-web/)
    web_root = os.path.abspath(os.path.join(os.path.dirname(args.out_json), ".."))
    assets_dir_abs = os.path.abspath(args.assets_dir)
    assets_rel_from_root = os.path.relpath(assets_dir_abs, web_root).replace("\\", "/")

    items: List[Dict[str, Any]] = []
    current_category = "未分类"

    def normalize_category(cat: str) -> str:
        c = (cat or "").strip()
        if c.startswith("当夏甜品"):
            return "当夏甜品"
        if c.startswith("当夏咸食"):
            return "当夏咸食|冷餐"
        if c.startswith("当夏水果"):
            return "当夏水果|茶饮"
        return c or "未分类"

    # Read all needed rows first
    max_col = ws.max_column
    max_row = ws.max_row

    def get_row(r: int) -> List[Any]:
        return [ws.cell(r, c).value for c in range(1, max_col + 1)]

    with zipfile.ZipFile(args.xlsx) as z:
        for r in range(1, max_row + 1):
            row = get_row(r)
            cat = is_category_row(row)
            if cat:
                current_category = normalize_category(cat)
                continue

            if not (isinstance(row[0], str) and row[0].strip() == "产品"):
                continue

            # Expect layout:
            # r: 产品
            # r+1: 参考图片 (cells contain DISPIMG formula IDs)
            # r+2: 单价（元）
            # r+3: 起订量
            img_row = get_row(r + 1) if r + 1 <= max_row else []
            price_row = get_row(r + 2) if r + 2 <= max_row else []
            min_row = get_row(r + 3) if r + 3 <= max_row else []

            for c in range(2, max_col + 1):
                name = ws.cell(r, c).value
                if not isinstance(name, str) or not name.strip():
                    continue
                name = name.strip()

                unit_price = ws.cell(r + 2, c).value if r + 2 <= max_row else None
                try:
                    unit_price_num = float(unit_price) if unit_price is not None and str(unit_price).strip() else None
                except Exception:
                    unit_price_num = None

                min_order = ws.cell(r + 3, c).value if r + 3 <= max_row else None
                min_order_str = str(min_order).strip() if min_order is not None else None

                disp = ws.cell(r + 1, c).value if r + 1 <= max_row else None
                img_id = parse_dispimg_id(disp)
                image_path = None

                media_path = None
                # 1) If there's an anchored image at that cell, use it
                media_path = cell_to_media.get((r + 1, c))
                # 2) Fallback: WPS DISPIMG mapping if present
                if media_path is None and img_id and img_id in id_to_media:
                    media_path = id_to_media[img_id]

                if media_path:
                    ext = os.path.splitext(media_path)[1].lstrip(".").lower() or "jpg"
                    filename = f"{current_category}_{slugify(name)}.{ext}"
                    full = os.path.join(args.assets_dir, filename)
                    if not os.path.exists(full):
                        data = z.read(media_path)
                        with open(full, "wb") as f:
                            f.write(data)
                    image_path = f"{assets_rel_from_root}/{filename}"

                items.append(
                    {
                        "category": normalize_category(current_category),
                        "name": name,
                        "unitPrice": unit_price_num,
                        "minOrder": min_order_str,
                        "minOrderNum": parse_min_order(min_order_str),
                        "image": image_path,
                    }
                )

    # de-dupe by (category,name)
    seen = set()
    dedup: List[Dict[str, Any]] = []
    for it in items:
        key = (it.get("category"), it.get("name"))
        if key in seen:
            continue
        seen.add(key)
        dedup.append(it)

    with open(args.out_json, "w", encoding="utf-8") as f:
        json.dump(dedup, f, ensure_ascii=False, indent=2)

    print(f"items: {len(dedup)}; images: {sum(1 for x in dedup if x.get('image'))}")


if __name__ == "__main__":
    main()
